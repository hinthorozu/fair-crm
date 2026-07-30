"""Missing-info filters and Excel export for the Customers screen."""

from datetime import UTC, datetime
from io import BytesIO
from uuid import UUID, uuid4

from openpyxl import load_workbook

from app.modules.customers.domain.value_objects import CustomerStatus, CustomerType
from app.modules.customers.infrastructure.persistence.communication_models import (
    CustomerEmailModel,
    CustomerPhoneModel,
    CustomerWebsiteModel,
)
from app.modules.customers.infrastructure.persistence.models import CustomerModel
from app.modules.participations.infrastructure.persistence.models import (
    CustomerFairParticipationModel,
)
from tests.conftest_helpers import pagination_from


def _create_fair(client, auth_headers, name: str) -> str:
    response = client.post(
        "/api/v1/fairs",
        headers=auth_headers,
        json={
            "name": name,
            "location": "Istanbul",
            "start_date": "2026-06-01",
            "end_date": "2026-06-03",
        },
    )
    assert response.status_code == 201
    return response.json()["id"]


def _seed_customer(
    db_session,
    organization_id,
    *,
    display_name: str,
    status: str = CustomerStatus.ACTIVE.value,
    phone: str | None = None,
    email: str | None = None,
    website: str | None = None,
):
    now = datetime.now(tz=UTC)
    customer_id = uuid4()
    customer = CustomerModel(
        id=customer_id,
        organization_id=organization_id,
        display_name=display_name,
        normalized_name=display_name.lower(),
        customer_type=CustomerType.EXHIBITOR.value,
        status=status,
        source="manual",
        created_at=now,
        updated_at=now,
    )
    db_session.add(customer)
    db_session.flush()
    if phone is not None:
        db_session.add(
            CustomerPhoneModel(
                id=uuid4(),
                organization_id=organization_id,
                customer_id=customer_id,
                phone=phone,
                is_primary=True,
                created_at=now,
            )
        )
    if email is not None:
        db_session.add(
            CustomerEmailModel(
                id=uuid4(),
                organization_id=organization_id,
                customer_id=customer_id,
                email=email,
                is_primary=True,
                created_at=now,
            )
        )
    if website is not None:
        db_session.add(
            CustomerWebsiteModel(
                id=uuid4(),
                organization_id=organization_id,
                customer_id=customer_id,
                website=website,
                is_primary=True,
                created_at=now,
            )
        )
    db_session.flush()
    return customer


def _link_fair(db_session, organization_id, customer_id, fair_id, *, deleted: bool = False):
    now = datetime.now(tz=UTC)
    db_session.add(
        CustomerFairParticipationModel(
            id=uuid4(),
            organization_id=organization_id,
            customer_id=customer_id,
            fair_id=UUID(str(fair_id)),
            participation_status="exhibitor",
            is_active=True,
            created_at=now,
            updated_at=now,
            deleted_at=now if deleted else None,
        )
    )
    db_session.flush()


def test_list_customers_missing_info_filters(client, auth_headers, db_session, organization_id):
    with_all = _seed_customer(
        db_session,
        organization_id,
        display_name="ABC Complete Co",
        phone="0212 111 1111",
        email="complete@example.com",
        website="https://complete.example",
    )
    no_website = _seed_customer(
        db_session,
        organization_id,
        display_name="ABC No Website Co",
        phone="0212 222 2222",
        email="noweb@example.com",
    )
    whitespace_website = _seed_customer(
        db_session,
        organization_id,
        display_name="ABC Whitespace Website Co",
        phone="0212 333 3333",
        email="spaceweb@example.com",
        website="   ",
    )
    no_phone = _seed_customer(
        db_session,
        organization_id,
        display_name="ABC No Phone Co",
        email="nophone@example.com",
        website="https://nophone.example",
    )
    no_email = _seed_customer(
        db_session,
        organization_id,
        display_name="ABC No Email Co",
        phone="0212 444 4444",
        website="https://noemail.example",
    )
    no_fair = _seed_customer(
        db_session,
        organization_id,
        display_name="ABC No Fair Co",
        phone="0212 555 5555",
        email="nofair@example.com",
        website="https://nofair.example",
    )
    with_fair = _seed_customer(
        db_session,
        organization_id,
        display_name="ABC With Fair Co",
        phone="0212 666 6666",
        email="withfair@example.com",
        website="https://withfair.example",
    )
    fair_id = _create_fair(client, auth_headers, "Missing Info Fair")
    _link_fair(db_session, organization_id, with_all.id, fair_id)
    _link_fair(db_session, organization_id, with_fair.id, fair_id)
    _link_fair(db_session, organization_id, no_fair.id, fair_id, deleted=True)
    db_session.commit()

    names = lambda response: {item["display_name"] for item in response.json()["items"]}

    no_website_res = client.get(
        "/api/v1/customers?search=ABC&status=active&missing_info=no_website&pageSize=100",
        headers=auth_headers,
    )
    assert no_website_res.status_code == 200
    assert names(no_website_res) == {
        "ABC No Website Co",
        "ABC Whitespace Website Co",
    }
    assert pagination_from(no_website_res.json())["totalItems"] == 2

    no_phone_res = client.get(
        "/api/v1/customers?search=ABC&status=active&missing_info=no_phone&pageSize=100",
        headers=auth_headers,
    )
    assert no_phone_res.status_code == 200
    assert names(no_phone_res) == {"ABC No Phone Co"}

    no_email_res = client.get(
        "/api/v1/customers?search=ABC&status=active&missing_info=no_email&pageSize=100",
        headers=auth_headers,
    )
    assert no_email_res.status_code == 200
    assert names(no_email_res) == {"ABC No Email Co"}

    no_fair_res = client.get(
        "/api/v1/customers?search=ABC&status=active&missing_info=no_fair&pageSize=100",
        headers=auth_headers,
    )
    assert no_fair_res.status_code == 200
    assert "ABC No Fair Co" in names(no_fair_res)
    assert "ABC With Fair Co" not in names(no_fair_res)
    assert "ABC Complete Co" not in names(no_fair_res)


def test_export_customers_respects_filters_and_format(
    client, auth_headers, db_session, organization_id
):
    multi = _seed_customer(
        db_session,
        organization_id,
        display_name="Export Multi Co",
        status=CustomerStatus.ACTIVE.value,
        website="https://export.example",
    )
    now = datetime.now(tz=UTC)
    db_session.add_all(
        [
            CustomerEmailModel(
                id=uuid4(),
                organization_id=organization_id,
                customer_id=multi.id,
                email="a@export.example",
                is_primary=True,
                created_at=now,
            ),
            CustomerEmailModel(
                id=uuid4(),
                organization_id=organization_id,
                customer_id=multi.id,
                email="b@export.example",
                is_primary=False,
                created_at=now,
            ),
            CustomerPhoneModel(
                id=uuid4(),
                organization_id=organization_id,
                customer_id=multi.id,
                phone="0212 777 0001",
                is_primary=True,
                created_at=now,
            ),
            CustomerPhoneModel(
                id=uuid4(),
                organization_id=organization_id,
                customer_id=multi.id,
                phone="0212 777 0002",
                is_primary=False,
                created_at=now,
            ),
        ]
    )
    no_website = _seed_customer(
        db_session,
        organization_id,
        display_name="Export No Website Co",
        status=CustomerStatus.ACTIVE.value,
        phone="0212 888 0000",
        email="noweb@export.example",
    )
    other = _seed_customer(
        db_session,
        organization_id,
        display_name="Other Active With Site",
        status=CustomerStatus.ACTIVE.value,
        website="https://other.example",
        phone="0212 999 0000",
        email="other@export.example",
    )
    fair_a = _create_fair(client, auth_headers, "Alpha Fair")
    fair_b = _create_fair(client, auth_headers, "Beta Fair")
    _link_fair(db_session, organization_id, multi.id, fair_a)
    _link_fair(db_session, organization_id, multi.id, fair_b)
    db_session.commit()

    list_res = client.get(
        "/api/v1/customers?search=Export&status=active&missing_info=no_website&pageSize=100",
        headers=auth_headers,
    )
    assert list_res.status_code == 200
    list_names = {item["display_name"] for item in list_res.json()["items"]}
    assert list_names == {"Export No Website Co"}
    list_total = pagination_from(list_res.json())["totalItems"]

    export_res = client.get(
        "/api/v1/customers/export?search=Export&status=active&missing_info=no_website",
        headers=auth_headers,
    )
    assert export_res.status_code == 200
    assert (
        export_res.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(filename=BytesIO(export_res.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "Customer UID"
    assert rows[0][1] == "Müşteri Adı"
    assert rows[0][-2:] == ("E-posta", "Telefon")
    assert rows[0][-3] == "Fuarlar"
    data_rows = rows[1:]
    assert len(data_rows) == list_total == 1
    assert data_rows[0][0] == str(no_website.id)
    assert data_rows[0][1] == "Export No Website Co"
    assert data_rows[0][-3] in ("", None)
    assert "noweb@export.example" in str(data_rows[0][-2])
    assert "0212 888 0000" in str(data_rows[0][-1])

    # Same customer UID must be stable across repeated exports.
    export_again = client.get(
        "/api/v1/customers/export?search=Export&status=active&missing_info=no_website",
        headers=auth_headers,
    )
    assert export_again.status_code == 200
    again_rows = list(
        load_workbook(filename=BytesIO(export_again.content)).active.iter_rows(values_only=True)
    )[1:]
    assert again_rows[0][0] == str(no_website.id)

    multi_export = client.get(
        "/api/v1/customers/export?search=Export+Multi&status=active",
        headers=auth_headers,
    )
    assert multi_export.status_code == 200
    multi_wb = load_workbook(filename=BytesIO(multi_export.content))
    multi_rows = list(multi_wb.active.iter_rows(values_only=True))[1:]
    assert len(multi_rows) == 1
    fairs_cell = str(multi_rows[0][-3])
    assert "Alpha Fair" in fairs_cell
    assert "Beta Fair" in fairs_cell
    emails_cell = str(multi_rows[0][-2])
    assert "a@export.example" in emails_cell
    assert "b@export.example" in emails_cell
    phones_cell = str(multi_rows[0][-1])
    assert "0212 777 0001" in phones_cell
    assert "0212 777 0002" in phones_cell
    assert multi_rows[0][0] == str(multi.id)
    assert multi_rows[0][1] == multi.display_name
    assert other.display_name not in {row[1] for row in multi_rows}


def test_export_deduplicates_fair_names_for_same_customer(
    client, auth_headers, db_session, organization_id
):
    customer = _seed_customer(
        db_session,
        organization_id,
        display_name="Dup Fair Export Co",
        status=CustomerStatus.ACTIVE.value,
        phone="0212 100 0000",
        email="dupfair@export.example",
        website="https://dupfair.example",
    )
    # Two distinct fair records with the same display name (common duplicate source),
    # plus a second unique fair.
    fair_a = _create_fair(client, auth_headers, "Frankfurt Book Fair 2025")
    fair_b = _create_fair(client, auth_headers, "Frankfurt Book Fair 2025")
    fair_c = _create_fair(client, auth_headers, "London Book Fair")
    _link_fair(db_session, organization_id, customer.id, fair_a)
    _link_fair(db_session, organization_id, customer.id, fair_b)
    _link_fair(db_session, organization_id, customer.id, fair_c)
    db_session.commit()

    list_res = client.get(
        "/api/v1/customers?search=Dup+Fair+Export&status=active&pageSize=100",
        headers=auth_headers,
    )
    assert list_res.status_code == 200
    list_total = pagination_from(list_res.json())["totalItems"]
    assert list_total == 1

    export_res = client.get(
        "/api/v1/customers/export?search=Dup+Fair+Export&status=active",
        headers=auth_headers,
    )
    assert export_res.status_code == 200
    rows = list(load_workbook(filename=BytesIO(export_res.content)).active.iter_rows(values_only=True))
    data_rows = rows[1:]
    assert len(data_rows) == list_total == 1
    fairs_cell = str(data_rows[0][-3])
    assert fairs_cell.count("Frankfurt Book Fair 2025") == 1
    assert "London Book Fair" in fairs_cell
    assert fairs_cell.count(",") == 1


def test_export_row_count_matches_list_total_for_combined_filters(
    client, auth_headers, db_session, organization_id
):
    _seed_customer(
        db_session,
        organization_id,
        display_name="Parity Match Co",
        status=CustomerStatus.ACTIVE.value,
        phone="0212 200 0001",
        email="parity@export.example",
    )
    _seed_customer(
        db_session,
        organization_id,
        display_name="Parity Match Co With Site",
        status=CustomerStatus.ACTIVE.value,
        phone="0212 200 0002",
        email="parity2@export.example",
        website="https://parity.example",
    )
    _seed_customer(
        db_session,
        organization_id,
        display_name="Parity Inactive Co",
        status=CustomerStatus.INACTIVE.value,
        phone="0212 200 0003",
        email="parity3@export.example",
    )
    db_session.commit()

    query = "search=Parity+Match&status=active&missing_info=no_website&customer_type=exhibitor"
    list_res = client.get(f"/api/v1/customers?{query}&pageSize=100", headers=auth_headers)
    assert list_res.status_code == 200
    list_total = pagination_from(list_res.json())["totalItems"]
    assert list_total == 1
    assert list_res.json()["items"][0]["display_name"] == "Parity Match Co"
    list_uid = list_res.json()["items"][0]["id"]

    export_res = client.get(f"/api/v1/customers/export?{query}", headers=auth_headers)
    assert export_res.status_code == 200
    data_rows = list(
        load_workbook(filename=BytesIO(export_res.content)).active.iter_rows(values_only=True)
    )[1:]
    assert len(data_rows) == list_total
    assert data_rows[0][0] == list_uid
    assert data_rows[0][1] == "Parity Match Co"
