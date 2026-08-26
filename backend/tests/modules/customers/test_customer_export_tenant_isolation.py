"""Adversarial tenant-isolation coverage for customer Excel export."""

from io import BytesIO
from uuid import UUID

from openpyxl import load_workbook

from app.modules.participations.infrastructure.persistence.models import (
    CustomerFairParticipationModel,
)


def _create_customer(client, headers, name: str) -> str:
    response = client.post(
        "/api/v1/customers",
        headers=headers,
        json={"display_name": name, "status": "active"},
    )
    assert response.status_code == 201, response.text
    return response.json()["id"]


def _create_fair(client, headers, name: str) -> str:
    response = client.post(
        "/api/v1/fairs",
        headers=headers,
        json={"name": name, "status": "planned"},
    )
    assert response.status_code == 201, response.text
    return response.json()["id"]


def _create_participation(client, headers, customer_id: str, fair_id: str) -> str:
    response = client.post(
        "/api/v1/fair-participations",
        headers=headers,
        json={"customer_id": customer_id, "fair_id": fair_id},
    )
    assert response.status_code == 201, response.text
    return response.json()["id"]


def _export_rows(response) -> list[dict[str, str | None]]:
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    sheet = workbook["customers"]
    values = list(sheet.iter_rows(values_only=True))
    headers = [str(value) for value in values[0]]
    return [dict(zip(headers, row, strict=True)) for row in values[1:]]


def test_customer_export_does_not_follow_cross_tenant_participation_or_fair(
    client,
    auth_headers,
    db_session,
    other_organization_id,
):
    owner_customer_id = _create_customer(client, auth_headers, "Export Owner Customer")
    owner_fair_id = _create_fair(client, auth_headers, "OWNER FAIR")
    _create_participation(client, auth_headers, owner_customer_id, owner_fair_id)

    owner_corrupt_fair_source_id = _create_fair(client, auth_headers, "Temporary Owner Fair")
    owner_corrupt_participation_id = _create_participation(
        client,
        auth_headers,
        owner_customer_id,
        owner_corrupt_fair_source_id,
    )

    other_headers = {**auth_headers, "X-Organization-Id": str(other_organization_id)}
    foreign_customer_id = _create_customer(client, other_headers, "Foreign Export Customer")
    foreign_fair_id = _create_fair(client, other_headers, "FOREIGN FAIR")
    foreign_participation_id = _create_participation(
        client,
        other_headers,
        foreign_customer_id,
        foreign_fair_id,
    )

    owner_corrupt_participation = db_session.get(
        CustomerFairParticipationModel,
        UUID(owner_corrupt_participation_id),
    )
    assert owner_corrupt_participation is not None
    owner_corrupt_participation.fair_id = UUID(foreign_fair_id)

    foreign_participation = db_session.get(
        CustomerFairParticipationModel,
        UUID(foreign_participation_id),
    )
    assert foreign_participation is not None
    foreign_participation.customer_id = UUID(owner_customer_id)
    db_session.commit()

    response = client.get("/api/v1/customers/export", headers=auth_headers)
    rows = _export_rows(response)

    owner_row = next(row for row in rows if row["Customer UID"] == owner_customer_id)
    assert owner_row["Fuarlar"] == "OWNER FAIR"
    assert "FOREIGN FAIR" not in str(owner_row["Fuarlar"])
    assert all(row["Customer UID"] != foreign_customer_id for row in rows)
