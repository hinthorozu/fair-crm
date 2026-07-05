"""Tests for scraper handoff Excel exporter."""

import json
from pathlib import Path

from openpyxl import load_workbook

from app.modules.scraper.exporters.scraper_excel_exporter import (
    EXCEL_COLUMNS,
    build_excel_rows,
    excel_columns_for_requested_fields,
    write_handoff_excel,
)
from app.modules.scraper.exporters.scraper_import_exporter import ScraperImportHandoff
from app.shared.canonical_import.scraper_mapper import scraper_handoff_to_canonical

_LIVE_HANDOFF_JSON = (
    Path(__file__).resolve().parents[3] / "data" / "scraper-handoff" / "11719501-5fcd-4dd4-8058-4154a86fbeeb.json"
)


def _handoff_from_canonical_json(doc: dict) -> ScraperImportHandoff:
    """Rebuild handoff the way normalizer stores social URLs on canonical rows."""
    canonical_rows: list[dict[str, str]] = []
    row_metadata: list[dict] = []
    for row in doc["rows"]:
        canonical: dict[str, str] = {"company_name": row["company_name"]}
        for field in (
            "website",
            "country",
            "city",
            "hall",
            "stand",
            "instagram_url",
            "facebook_url",
            "linkedin_url",
            "youtube_url",
        ):
            value = row.get(field)
            if value:
                canonical[field] = value
        if row.get("emails"):
            canonical["email"] = row["emails"][0]
        if row.get("phones"):
            canonical["phone"] = row["phones"][0]
        canonical_rows.append(canonical)
        row_metadata.append(dict(row.get("raw") or {}))
    return ScraperImportHandoff(
        canonical_rows=canonical_rows,
        row_metadata=row_metadata,
        metadata={"source_site": doc["source"]["adapter_key"]},
    )


def test_write_handoff_excel_writes_expected_headers(tmp_path: Path):
    handoff = ScraperImportHandoff(
        canonical_rows=[
            {
                "company_name": "Acme",
                "country": "Türkiye",
                "hall": "1",
                "stand": "A1",
            }
        ],
        metadata={"fair_name": "Foodist Expo"},
        row_metadata=[
            {
                "source_url": "https://example.test/brand/acme",
                "category": "Gıda",
                "detail_scraped": True,
                "website_valid": True,
            }
        ],
    )

    path = write_handoff_excel(handoff, str(tmp_path / "companies.xlsx"))
    workbook = load_workbook(path)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == list(EXCEL_COLUMNS)


def test_write_handoff_excel_filters_columns_by_requested_fields(tmp_path: Path):
    handoff = ScraperImportHandoff(
        canonical_rows=[
            {
                "company_name": "Acme",
                "phone": "555",
                "email": "a@b.com",
                "website": "https://acme.test",
            }
        ],
        row_metadata=[{}],
    )
    requested = ["customerName", "phone", "email"]
    columns = excel_columns_for_requested_fields(requested)
    assert columns == ("company_name", "phone", "email")

    path = write_handoff_excel(
        handoff,
        str(tmp_path / "filtered.xlsx"),
        requested_fields=requested,
    )
    workbook = load_workbook(path)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == list(columns)
    assert [cell.value for cell in sheet[2]] == ["Acme", "555", "a@b.com"]


def test_write_handoff_excel_includes_social_urls_from_canonical_rows(tmp_path: Path):
    handoff = ScraperImportHandoff(
        canonical_rows=[
            {
                "company_name": "Social Co",
                "instagram_url": "https://instagram.com/socialco",
                "facebook_url": "https://facebook.com/socialco",
                "linkedin_url": "https://linkedin.com/company/socialco",
                "youtube_url": "https://youtube.com/@socialco",
            }
        ],
        row_metadata=[{}],
    )
    requested = ["customerName", "instagram", "facebook", "linkedin", "youtube"]
    path = write_handoff_excel(
        handoff,
        str(tmp_path / "social.xlsx"),
        requested_fields=requested,
    )
    workbook = load_workbook(path)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == [
        "company_name",
        "instagram_url",
        "facebook_url",
        "linkedin_url",
        "youtube_url",
    ]
    assert [cell.value for cell in sheet[2]] == [
        "Social Co",
        "https://instagram.com/socialco",
        "https://facebook.com/socialco",
        "https://linkedin.com/company/socialco",
        "https://youtube.com/@socialco",
    ]


def test_write_handoff_excel_resolves_short_social_keys_from_metadata(tmp_path: Path):
    handoff = ScraperImportHandoff(
        canonical_rows=[{"company_name": "Legacy Meta Co"}],
        row_metadata=[
            {
                "instagram": "instagram.com/legacy",
                "facebook": "facebook.com/legacy",
            }
        ],
    )
    requested = ["customerName", "instagram", "facebook"]
    path = write_handoff_excel(
        handoff,
        str(tmp_path / "social-short.xlsx"),
        requested_fields=requested,
    )
    workbook = load_workbook(path)
    sheet = workbook.active
    assert [cell.value for cell in sheet[2]] == [
        "Legacy Meta Co",
        "https://instagram.com/legacy",
        "https://facebook.com/legacy",
    ]


def test_excel_social_urls_match_json_from_live_handoff_11719501(tmp_path: Path):
    if not _LIVE_HANDOFF_JSON.is_file():
        return

    doc = json.loads(_LIVE_HANDOFF_JSON.read_text(encoding="utf-8"))
    handoff = _handoff_from_canonical_json(doc)
    adapter_key = doc["source"]["adapter_key"]
    run_id = doc["source"].get("run_id")
    source_url = doc["source"].get("source_url")

    json_document = scraper_handoff_to_canonical(
        handoff,
        adapter_key=adapter_key,
        run_id=run_id,
        source_url=source_url,
    )
    excel_rows = build_excel_rows(
        handoff,
        adapter_key=adapter_key,
        run_id=run_id,
        source_url=source_url,
    )
    assert len(excel_rows) == len(json_document.rows)

    social_columns = ("instagram_url", "facebook_url", "linkedin_url", "youtube_url")
    mismatches: list[str] = []
    for index, (json_row, excel_row) in enumerate(zip(json_document.rows, excel_rows, strict=True)):
        for column in social_columns:
            json_value = getattr(json_row, column) or ""
            excel_value = excel_row.get(column, "")
            if json_value != excel_value:
                mismatches.append(
                    f"row {index + 1} {json_row.company_name!r} {column}: json={json_value!r} excel={excel_value!r}"
                )
    assert not mismatches, "\n".join(mismatches[:10])

    requested = ["customerName", "instagram", "facebook", "linkedin", "youtube"]
    path = write_handoff_excel(
        handoff,
        str(tmp_path / "live-social.xlsx"),
        requested_fields=requested,
        adapter_key=adapter_key,
        run_id=run_id,
        source_url=source_url,
    )
    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    instagram_idx = headers.index("instagram_url")
    json_with_instagram = [
        (index, row)
        for index, row in enumerate(json_document.rows, start=2)
        if row.instagram_url
    ]
    assert json_with_instagram, "fixture JSON should contain instagram_url rows"
    for sheet_row, json_row in zip(
        sheet.iter_rows(min_row=2, values_only=True),
        json_document.rows,
        strict=True,
    ):
        assert (sheet_row[instagram_idx] or "") == (json_row.instagram_url or "")
