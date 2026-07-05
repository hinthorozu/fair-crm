"""Manual column mapping for Universal Import Engine."""

from collections.abc import Iterator
from typing import Any

from openpyxl.utils import get_column_letter

from app.modules.imports.domain.exceptions import InvalidColumnMappingError
from app.modules.imports.domain.services.header_mapping import map_header_to_field, normalize_header
from app.modules.imports.domain.value_objects import ExcelHeaderMode
from app.shared.import_output_fields import (
    GRID_MAPPING_FIELDS,
    WIZARD_MAPPING_FIELDS,
)

REJECTED_MAPPING_FIELDS = frozenset({"fair_name"})

MAX_MAPPING_SAMPLE_ROWS = 10
MAX_GRID_PREVIEW_ROWS = 50


def _is_empty_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _data_rows_for_preview(
    rows: list[list[Any]],
    mode: ExcelHeaderMode,
    header_row_index: int | None,
    *,
    max_sample_rows: int = MAX_MAPPING_SAMPLE_ROWS,
) -> list[list[Any]]:
    if mode == ExcelHeaderMode.NO_HEADER:
        return rows[:max_sample_rows]
    start = int(header_row_index or 0) + 1
    return rows[start : start + max_sample_rows]


def header_mode_to_has_header(mode: ExcelHeaderMode) -> bool:
    return mode != ExcelHeaderMode.NO_HEADER


def resolve_header_mode(
    *,
    header_mode: ExcelHeaderMode | None = None,
    has_header_row: bool | None = None,
) -> ExcelHeaderMode:
    if header_mode is not None:
        return header_mode
    if has_header_row is False:
        return ExcelHeaderMode.NO_HEADER
    return ExcelHeaderMode.FIRST_ROW_HEADER


def resolve_header_row_index(
    mode: ExcelHeaderMode,
    *,
    header_row_index: int | None = None,
) -> int | None:
    if mode == ExcelHeaderMode.NO_HEADER:
        return None
    if mode == ExcelHeaderMode.MANUAL_HEADER_ROW:
        if header_row_index is None or header_row_index < 0:
            raise InvalidColumnMappingError("header_row_index is required for manual_header_row mode")
        return header_row_index
    return 0


def _headers_for_mode(
    raw_preview: dict[str, Any],
    mode: ExcelHeaderMode,
    header_row_index: int | None,
) -> list[Any]:
    rows: list[list[Any]] = raw_preview.get("rows") or []
    if mode == ExcelHeaderMode.NO_HEADER:
        return []
    index = header_row_index if header_row_index is not None else 0
    if index >= len(rows):
        return []
    return [str(v) if v is not None else None for v in rows[index]]


def suggest_column_mapping(
    raw_preview: dict[str, Any],
    *,
    has_header_row: bool | None = None,
    header_mode: ExcelHeaderMode | None = None,
) -> dict[str, Any]:
    mode = resolve_header_mode(header_mode=header_mode, has_header_row=has_header_row)
    header_row_index = resolve_header_row_index(mode)
    headers = _headers_for_mode(raw_preview, mode, header_row_index)

    mappings: dict[str, dict[str, Any]] = {}
    if mode != ExcelHeaderMode.NO_HEADER and headers:
        for index, header in enumerate(headers):
            if header is None:
                continue
            field = map_header_to_field(str(header))
            if field and field in GRID_MAPPING_FIELDS and field not in mappings:
                mappings[field] = {"type": "column_index", "value": index}

    return {
        "header_mode": mode.value,
        "has_header_row": header_mode_to_has_header(mode),
        "header_row_index": header_row_index,
        "mappings": mappings,
    }


def validate_column_mapping(mapping_config: dict[str, Any]) -> None:
    mappings = mapping_config.get("mappings") or {}
    if not isinstance(mappings, dict):
        raise InvalidColumnMappingError("Invalid mapping format")

    for field in mappings:
        if field in REJECTED_MAPPING_FIELDS:
            raise InvalidColumnMappingError(f"Field '{field}' is not supported")
        if field not in WIZARD_MAPPING_FIELDS:
            raise InvalidColumnMappingError(f"Unknown mapping field: {field}")

    if "company_name" not in mappings:
        raise InvalidColumnMappingError("company_name mapping is required")

    seen_columns: set[int] = set()
    for field, spec in mappings.items():
        if not isinstance(spec, dict):
            raise InvalidColumnMappingError(f"Invalid mapping spec for {field}")
        if spec.get("type") != "column_index":
            raise InvalidColumnMappingError(f"Unsupported mapping type for {field}")
        try:
            col_index = int(spec["value"])
        except (KeyError, TypeError, ValueError) as exc:
            raise InvalidColumnMappingError(f"Invalid column index for {field}") from exc
        if col_index in seen_columns:
            raise InvalidColumnMappingError(f"Duplicate column mapping at index {col_index}")
        seen_columns.add(col_index)
        _ = field


def _iter_data_rows(
    rows: list[list[Any]],
    mapping_config: dict[str, Any],
) -> Iterator[list[Any]]:
    mode = resolve_header_mode(
        header_mode=ExcelHeaderMode(mapping_config["header_mode"])
        if mapping_config.get("header_mode")
        else None,
        has_header_row=mapping_config.get("has_header_row"),
    )
    header_row_index = mapping_config.get("header_row_index")
    if header_row_index is None and mode != ExcelHeaderMode.NO_HEADER:
        header_row_index = 0

    if mode == ExcelHeaderMode.NO_HEADER:
        yield from rows
        return

    start = int(header_row_index or 0) + 1
    if start <= len(rows):
        yield from rows[start:]


def iter_mapped_rows(
    raw_preview: dict[str, Any],
    mapping_config: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    """Yield mapped row dicts one at a time (memory-friendly analyze path)."""
    validate_column_mapping(mapping_config)
    rows: list[list[Any]] = raw_preview.get("rows") or []
    mappings: dict[str, dict[str, Any]] = mapping_config.get("mappings") or {}

    for row in _iter_data_rows(rows, mapping_config):
        raw: dict[str, Any] = {}
        for field, spec in mappings.items():
            index = int(spec["value"])
            if index < len(row):
                value = row[index]
                if value is not None and str(value).strip():
                    raw[field] = value
        if raw:
            yield raw


def apply_column_mapping(
    raw_preview: dict[str, Any],
    mapping_config: dict[str, Any],
) -> list[dict[str, Any]]:
    return list(iter_mapped_rows(raw_preview, mapping_config))


def build_mapping_preview_columns(
    raw_preview: dict[str, Any],
    *,
    has_header_row: bool | None = None,
    header_mode: ExcelHeaderMode | None = None,
    header_row_index: int | None = None,
    max_sample_rows: int = MAX_MAPPING_SAMPLE_ROWS,
) -> list[dict[str, Any]]:
    """Universal Import Mapping Preview — per-column header + sample values (max 10 rows)."""
    mode = resolve_header_mode(header_mode=header_mode, has_header_row=has_header_row)
    resolved_index = resolve_header_row_index(mode, header_row_index=header_row_index)
    headers = _headers_for_mode(raw_preview, mode, resolved_index)
    rows: list[list[Any]] = raw_preview.get("rows") or []
    columns_meta: list[dict[str, Any]] = raw_preview.get("columns") or []
    max_cols = len(columns_meta) if columns_meta else (max(len(r) for r in rows) if rows else 0)
    data_rows = _data_rows_for_preview(rows, mode, resolved_index, max_sample_rows=max_sample_rows)

    result: list[dict[str, Any]] = []
    for index in range(max_cols):
        letter = columns_meta[index]["letter"] if index < len(columns_meta) else get_column_letter(index + 1)
        header: str | None = None
        if mode != ExcelHeaderMode.NO_HEADER and index < len(headers) and headers[index]:
            header = str(headers[index])

        samples: list[Any] = []
        empty_count = 0
        for row in data_rows:
            value = row[index] if index < len(row) else None
            if _is_empty_cell(value):
                samples.append(None)
                empty_count += 1
            else:
                samples.append(value)

        filled_count = len(samples) - empty_count
        first_value = next((s for s in samples if s is not None), None)

        result.append(
            {
                "key": letter,
                "index": index,
                "letter": letter,
                "header": header,
                "samples": samples,
                "stats": {
                    "total": len(samples),
                    "empty": empty_count,
                    "filled": filled_count,
                    "first_value": str(first_value) if first_value is not None else None,
                },
            }
        )
    return result


def build_mapping_field_options(
    raw_preview: dict[str, Any],
    *,
    has_header_row: bool | None = None,
    header_mode: ExcelHeaderMode | None = None,
    header_row_index: int | None = None,
) -> list[dict[str, Any]]:
    preview_columns = build_mapping_preview_columns(
        raw_preview,
        has_header_row=has_header_row,
        header_mode=header_mode,
        header_row_index=header_row_index,
    )
    mode = resolve_header_mode(header_mode=header_mode, has_header_row=has_header_row)
    options: list[dict[str, Any]] = []

    for col in preview_columns:
        index = col["index"]
        letter = col["letter"]
        header = col.get("header")
        sample_values = col.get("samples") or []
        if mode == ExcelHeaderMode.NO_HEADER:
            samples = ", ".join(str(v) for v in sample_values[:3] if v is not None)
            label = f"Column {letter}"
            if samples:
                label = f"Column {letter} ({samples})"
        elif header:
            label = f"{header} ({letter})"
        else:
            label = f"Column {letter}"
        options.append(
            {
                "index": index,
                "letter": letter,
                "label": label,
                "sample_values": sample_values,
            }
        )
    return options


def build_excel_grid_preview(
    raw_preview: dict[str, Any],
    *,
    has_header_row: bool | None = None,
    header_mode: ExcelHeaderMode | None = None,
    header_row_index: int | None = None,
    max_rows: int = MAX_GRID_PREVIEW_ROWS,
) -> dict[str, Any]:
    """Excel-like grid for column mapping UI (header row + data rows)."""
    mode = resolve_header_mode(header_mode=header_mode, has_header_row=has_header_row)
    resolved_index = resolve_header_row_index(mode, header_row_index=header_row_index)
    rows: list[list[Any]] = raw_preview.get("rows") or []
    columns_meta: list[dict[str, Any]] = raw_preview.get("columns") or []
    max_cols = len(columns_meta) if columns_meta else (max(len(r) for r in rows) if rows else 0)

    if mode == ExcelHeaderMode.NO_HEADER:
        data_start = 0
        header_cells: list[str | None] = [
            columns_meta[i]["letter"] if i < len(columns_meta) else get_column_letter(i + 1)
            for i in range(max_cols)
        ]
    else:
        data_start = int(resolved_index or 0) + 1
        header_row = rows[resolved_index or 0] if rows else []
        header_cells = [
            str(header_row[i]) if i < len(header_row) and header_row[i] is not None else None
            for i in range(max_cols)
        ]

    data_rows = rows[data_start : data_start + max_rows]
    total_data_rows = max(len(rows) - data_start, 0) if mode != ExcelHeaderMode.NO_HEADER else len(rows)

    columns: list[dict[str, Any]] = []
    for index in range(max_cols):
        letter = columns_meta[index]["letter"] if index < len(columns_meta) else get_column_letter(index + 1)
        columns.append(
            {
                "index": index,
                "letter": letter,
                "header": header_cells[index] if index < len(header_cells) else None,
            }
        )

    return {
        "columns": columns,
        "rows": data_rows,
        "total_data_rows": total_data_rows,
        "preview_row_count": len(data_rows),
    }

