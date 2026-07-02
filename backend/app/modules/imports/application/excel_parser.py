from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.core.config import get_settings
from app.modules.imports.domain.exceptions import InvalidImportFileError
from app.modules.imports.domain.import_limits import ImportLimits
from app.modules.imports.domain.services.header_mapping import map_header_to_field


def parse_xlsx_rows(
    file_content: bytes,
    *,
    limits: ImportLimits | None = None,
) -> list[dict[str, Any]]:
    resolved_limits = limits or ImportLimits.from_settings(get_settings())
    resolved_limits.validate_file_size(len(file_content))

    try:
        workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    except Exception as exc:
        raise InvalidImportFileError("Invalid xlsx file") from exc

    sheet_count = len(workbook.worksheets)
    resolved_limits.validate_sheet_count(sheet_count)

    sheet = workbook.active
    if sheet is None:
        workbook.close()
        raise InvalidImportFileError("Worksheet is empty")

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        workbook.close()
        raise InvalidImportFileError("Worksheet has no header row") from exc

    field_indexes: dict[int, str] = {}
    for index, header in enumerate(header_row):
        if header is None:
            continue
        field = map_header_to_field(str(header))
        if field:
            field_indexes[index] = field

    if "company_name" not in field_indexes.values():
        workbook.close()
        raise InvalidImportFileError("Required column company_name (Firma Adı) not found")

    if len(field_indexes) > resolved_limits.max_columns:
        workbook.close()
        resolved_limits.validate_column_count(len(field_indexes))

    parsed_rows: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        if len(row) > resolved_limits.max_columns:
            workbook.close()
            resolved_limits.validate_column_count(len(row))
        raw: dict[str, Any] = {}
        for index, field in field_indexes.items():
            if index < len(row):
                value = row[index]
                if value is not None and str(value).strip():
                    raw[field] = value
        if raw:
            parsed_rows.append(raw)
            if len(parsed_rows) > resolved_limits.max_rows:
                workbook.close()
                resolved_limits.validate_row_count(len(parsed_rows))

    workbook.close()
    return parsed_rows
