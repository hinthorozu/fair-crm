"""Parse Excel into raw grid for wizard upload (no CRM mapping)."""

from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.core.config import Settings, get_settings
from app.modules.imports.domain.exceptions import InvalidImportFileError
from app.modules.imports.domain.import_limits import ImportLimits


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _extract_sheet_rows(sheet, *, limits: ImportLimits) -> list[list[Any]]:
    all_rows: list[list[Any]] = []
    max_row_width = 0
    for row in sheet.iter_rows(values_only=True):
        if row is None:
            continue
        cells = [_cell_value(cell) for cell in row]
        if all(c is None for c in cells):
            continue
        row_width = len(cells)
        while row_width > 0 and cells[row_width - 1] is None:
            row_width -= 1
        if row_width > limits.max_columns:
            limits.validate_column_count(row_width)
        max_row_width = max(max_row_width, row_width)
        all_rows.append(cells)
        if len(all_rows) > limits.max_rows:
            limits.validate_row_count(len(all_rows))
    if max_row_width > limits.max_columns:
        limits.validate_column_count(max_row_width)
    return all_rows


def _build_preview_from_rows(
    sheet_name: str,
    all_rows: list[list[Any]],
    *,
    limits: ImportLimits,
) -> dict[str, Any]:
    if not all_rows:
        raise InvalidImportFileError("File is empty or unreadable")

    max_cols = max(len(r) for r in all_rows)
    limits.validate_column_count(max_cols)

    normalized_rows: list[list[Any]] = []
    for row in all_rows:
        padded = list(row) + [None] * (max_cols - len(row))
        normalized_rows.append(padded[:max_cols])

    sample_cap = limits.mapping_sample_rows
    columns: list[dict[str, Any]] = []
    for index in range(max_cols):
        sample_values: list[Any] = []
        for row in normalized_rows[:sample_cap]:
            if index < len(row) and row[index] is not None:
                sample_values.append(row[index])
                if len(sample_values) >= sample_cap:
                    break
        columns.append(
            {
                "index": index,
                "letter": get_column_letter(index + 1),
                "sample_values": sample_values,
            }
        )

    first_row = normalized_rows[0]
    detected_headers = [str(v) if v is not None else None for v in first_row]

    return {
        "sheet_name": sheet_name,
        "rows": normalized_rows,
        "columns": columns,
        "detected_headers": detected_headers,
        "total_rows": len(normalized_rows),
    }


def parse_xlsx_raw(
    file_content: bytes,
    *,
    sheet_name: str | None = None,
    limits: ImportLimits | None = None,
    settings: Settings | None = None,
) -> dict[str, Any]:
    resolved_limits = limits or ImportLimits.from_settings(settings or get_settings())
    resolved_limits.validate_file_size(len(file_content))

    try:
        workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    except Exception as exc:
        raise InvalidImportFileError("Invalid xlsx file") from exc

    available_sheets = [ws.title for ws in workbook.worksheets]
    if not available_sheets:
        workbook.close()
        raise InvalidImportFileError("Workbook has no worksheets")

    resolved_limits.validate_sheet_count(len(available_sheets))

    target_name = sheet_name or available_sheets[0]
    if target_name not in available_sheets:
        workbook.close()
        raise InvalidImportFileError(f"Sheet '{target_name}' not found")

    sheet = workbook[target_name]
    if sheet is None:
        workbook.close()
        raise InvalidImportFileError("Worksheet is empty")

    all_rows = _extract_sheet_rows(sheet, limits=resolved_limits)
    workbook.close()

    preview = _build_preview_from_rows(sheet.title, all_rows, limits=resolved_limits)
    preview["available_sheets"] = available_sheets
    return preview
