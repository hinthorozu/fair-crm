"""Extract recipient rows from Excel workbooks for bulk-email preview/send."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

from app.modules.imports.domain.exceptions import InvalidImportFileError


@dataclass(frozen=True)
class ExcelRecipientRow:
    """One Excel data row: col1 display name + col2 raw email (unvalidated)."""

    display_name: str
    email: str


def extract_excel_recipient_rows(content: bytes) -> list[ExcelRecipientRow]:
    """Read col1 (display name) + col2 (email) from every sheet row.

    There is no bulk-email Excel header contract: every non-empty row is returned.
    A header like ``E-posta`` fails later validation/skip like any other invalid email.
    """
    if not content:
        raise InvalidImportFileError("Only .xlsx files are supported")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl raises varied parse errors
        raise InvalidImportFileError("Excel dosyası okunamadı") from exc

    rows: list[ExcelRecipientRow] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                if row is None:
                    continue
                cells = list(row)
                name_raw = cells[0] if len(cells) > 0 else None
                email_raw = cells[1] if len(cells) > 1 else None
                display_name = "" if name_raw is None else str(name_raw).strip()
                email = "" if email_raw is None else str(email_raw).strip()
                if not display_name and not email:
                    continue
                rows.append(ExcelRecipientRow(display_name=display_name, email=email))
    finally:
        workbook.close()

    return rows


def extract_email_tokens_from_xlsx(content: bytes) -> list[str]:
    """Backward-compatible: return raw email column values only."""
    return [row.email for row in extract_excel_recipient_rows(content) if row.email]
