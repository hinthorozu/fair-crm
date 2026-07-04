import json
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

backend = Path(__file__).resolve().parents[1]
json_path = backend / "foodist_handoff.json"
xlsx_path = backend / "foodist_companies.xlsx"
log_path = backend / "foodist_scraper_run.log"

summary_lines: list[str] = []

# JSON analysis
if json_path.exists():
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    rows = payload.get("canonical_rows", [])
    names = [r.get("company_name", "") for r in rows]
    empty_names = sum(1 for n in names if not str(n).strip())
    name_counts = Counter(names)
    duplicates = {k: v for k, v in name_counts.items() if v > 1 and k}
    email_filled = sum(1 for r in rows if str(r.get("email") or "").strip())
    website_filled = sum(1 for r in rows if str(r.get("website") or "").strip())
    phone_filled = sum(1 for r in rows if str(r.get("phone") or "").strip())
    summary_lines.extend(
        [
            f"JSON exists: yes ({json_path})",
            f"Total records (canonical_rows): {len(rows)}",
            f"Email filled: {email_filled}",
            f"Website filled: {website_filled}",
            f"Phone filled: {phone_filled}",
            f"Empty company_name: {empty_names}",
            f"Duplicate company_name count: {len(duplicates)}",
        ]
    )
    if duplicates:
        summary_lines.append("Duplicate examples:")
        for name, count in list(duplicates.items())[:5]:
            summary_lines.append(f"  - {name!r}: {count}")
else:
    summary_lines.append("JSON exists: no")

# Excel analysis
if xlsx_path.exists():
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    summary_lines.append(f"Excel exists: yes ({xlsx_path})")
    summary_lines.append(f"Excel rows (excl header): {len(data_rows)}")
    summary_lines.append(f"Excel columns: {headers}")
    wb.close()
else:
    summary_lines.append("Excel exists: no")

# Log errors
if log_path.exists():
    log_text = log_path.read_text(encoding="utf-8", errors="replace")
    error_lines = [line for line in log_text.splitlines() if "ERROR" in line or "Traceback" in line or "failed" in line.lower()]
    warning_lines = [line for line in log_text.splitlines() if "WARNING" in line]
    summary_lines.append(f"Log exists: yes")
    summary_lines.append(f"Log ERROR/Traceback lines: {len(error_lines)}")
    summary_lines.append(f"Log WARNING lines: {len(warning_lines)}")
    if error_lines[:10]:
        summary_lines.append("Sample errors:")
        summary_lines.extend(error_lines[:10])
else:
    summary_lines.append("Log exists: no")

Path(backend / "foodist_summary.txt").write_text("\n".join(summary_lines), encoding="utf-8")
