import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

backend = Path(__file__).resolve().parents[1]
json_path = backend / "foodist_handoff.json"
xlsx_path = backend / "foodist_companies.xlsx"

payload = json.loads(json_path.read_text(encoding="utf-8"))
rows = payload["canonical_rows"]
meta = payload.get("row_metadata", [])

total = len(rows)
names = [r.get("company_name", "") for r in rows]
countries = [r.get("country", "") for r in rows]
websites = [r.get("website", "") for r in rows]
halls = [r.get("hall", "") for r in rows]
stands = [r.get("stand", "") for r in rows]

country_in_name = sum(
    1
    for name in names
    if any(token in (name or "") for token in ("Türkiye", "Turkiye", "Almanya", "Çin", "Markalar"))
)
website_filled = sum(1 for w in websites if w)
bad_websites = sum(
    1
    for w in websites
    if w and any(bad in w.lower() for bad in ("googleapis", "gstatic", "foodistexpo", "googletagmanager", "instagram", "facebook"))
)
website_valid_true = sum(1 for m in meta if m.get("website_valid") is True)
website_valid_false = sum(1 for m in meta if m.get("website_valid") is False)
detail_scraped = sum(1 for m in meta if m.get("detail_scraped") is True)
hall_stand_ok = sum(1 for r in rows if r.get("hall") and r.get("stand"))

lines = [
    f"total={total}",
    f"country_populated={sum(1 for c in countries if c)}",
    f"country_in_company_name={country_in_name}",
    f"company_name_nonempty={sum(1 for n in names if str(n).strip())}",
    f"duplicate_names={sum(1 for k,v in Counter(names).items() if v>1 and k)}",
    f"website_filled={website_filled}",
    f"bad_website_count={bad_websites}",
    f"website_valid_true={website_valid_true}",
    f"website_valid_false={website_valid_false}",
    f"detail_scraped={detail_scraped}",
    f"hall_and_stand_filled={hall_stand_ok}",
    f"top_websites={Counter(websites).most_common(5)}",
]

if xlsx_path.exists():
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    lines.append(f"excel_headers={headers}")
    wb.close()

Path(backend / "foodist_quality_v1.txt").write_text("\n".join(lines), encoding="utf-8")
